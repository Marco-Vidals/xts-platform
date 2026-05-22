"""Genera Excel con la estructura completa de la BD XTS para presentacion."""
import pyodbc, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DB = (
    "DRIVER={ODBC Driver 17 for SQL Server};SERVER=100.70.216.12;"
    "DATABASE=XTS;UID=sa;PWD=XTS_operations.XiiX;"
)
conn = pyodbc.connect(DB)

# ── Esquema completo ──────────────────────────────────────────────────────────
df_schema = pd.read_sql("""
SELECT
    t.TABLE_SCHEMA,
    t.TABLE_NAME,
    c.COLUMN_NAME,
    c.DATA_TYPE,
    c.CHARACTER_MAXIMUM_LENGTH,
    c.IS_NULLABLE,
    CASE WHEN pk.COLUMN_NAME IS NOT NULL THEN 'SI' ELSE '' END AS IS_PK
FROM INFORMATION_SCHEMA.TABLES t
JOIN INFORMATION_SCHEMA.COLUMNS c
    ON t.TABLE_NAME = c.TABLE_NAME AND t.TABLE_SCHEMA = c.TABLE_SCHEMA
LEFT JOIN (
    SELECT ku.TABLE_SCHEMA, ku.TABLE_NAME, ku.COLUMN_NAME
    FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
    JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE ku
        ON tc.CONSTRAINT_NAME = ku.CONSTRAINT_NAME
    WHERE tc.CONSTRAINT_TYPE = 'PRIMARY KEY'
) pk ON pk.TABLE_SCHEMA = c.TABLE_SCHEMA
     AND pk.TABLE_NAME = c.TABLE_NAME
     AND pk.COLUMN_NAME = c.COLUMN_NAME
WHERE t.TABLE_TYPE = 'BASE TABLE'
ORDER BY t.TABLE_SCHEMA, t.TABLE_NAME, c.ORDINAL_POSITION
""", conn)

# ── Conteo de filas ───────────────────────────────────────────────────────────
row_counts = {}
for _, row in df_schema[["TABLE_SCHEMA","TABLE_NAME"]].drop_duplicates().iterrows():
    key = f"{row.TABLE_SCHEMA}.{row.TABLE_NAME}"
    try:
        cnt = pd.read_sql(
            f"SELECT COUNT(*) AS cnt FROM [{row.TABLE_SCHEMA}].[{row.TABLE_NAME}]",
            conn
        ).iloc[0, 0]
        row_counts[key] = cnt
    except Exception:
        row_counts[key] = "N/A"

# ── Rango de fechas ───────────────────────────────────────────────────────────
date_ranges = {}
for tbl in ["DATOS_ERCOT", "DATOS_CAISO", "DATOS_PML", "GTM"]:
    try:
        r = pd.read_sql(
            f"SELECT MIN(fecha) AS mn, MAX(fecha) AS mx FROM dbo.[{tbl}]", conn
        )
        date_ranges[tbl] = (str(r.iloc[0,0])[:10], str(r.iloc[0,1])[:10])
    except Exception:
        pass

conn.close()

# ── Estilos ───────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="002B4A")
ALT_FILL   = PatternFill("solid", fgColor="EBF3FA")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")

HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
TITLE_FONT = Font(bold=True, color="002B4A", name="Calibri", size=14)
BOLD_FONT  = Font(bold=True, name="Calibri", size=10)
NORM_FONT  = Font(name="Calibri", size=10)
ITALIC     = Font(italic=True, color="666666", name="Calibri", size=9)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ── Hoja 1: Resumen de Tablas ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Resumen Tablas"

ws1.merge_cells("A1:G1")
ws1["A1"] = "XTS Platform — Estructura de Base de Datos"
ws1["A1"].font = TITLE_FONT
ws1["A1"].alignment = CENTER
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:G2")
ws1["A2"] = "Servidor: 100.70.216.12  |  Base de datos: XTS  |  Actualizado: 2026-04-06"
ws1["A2"].font = ITALIC
ws1["A2"].alignment = CENTER
ws1.row_dimensions[2].height = 18

for col, h in enumerate(
    ["Schema", "Tabla", "Descripcion", "Filas", "Fecha Inicio", "Fecha Fin", "Frecuencia"],
    start=1,
):
    c = ws1.cell(row=4, column=col, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
ws1.row_dimensions[4].height = 22

TABLE_META = {
    "DATOS_ERCOT":            ("dbo",     "Precios horarios de energia ERCOT (Texas)", "Horaria"),
    "DATOS_CAISO":            ("dbo",     "Precios horarios de energia CAISO (California)", "Horaria"),
    "DATOS_PML":              ("dbo",     "Precios Marginales Locales CENACE por nodo", "Horaria"),
    "GTM":                    ("dbo",     "Precios Guatemala (PPOE forecast, POE real, LBR)", "Horaria"),
    "CAPACIDADES":            ("dbo",     "Capacidades maximas de transporte entre nodos", "Por evento"),
    "DEF_EXC":                ("dbo",     "Deficits y excedentes por nodo y hora", "Horaria"),
    "ESTADOS_DE_CUENTA_XML":  ("dbo",     "Estados de cuenta diarios BCA/SIN (ECD)", "Diaria"),
    "IMPEXP":                 ("dbo",     "Importaciones y exportaciones CENACE", "Horaria"),
    "TIPO_CAMBIO":            ("dbo",     "Tipo de cambio FIX Banxico (MXN/USD)", "Diaria"),
    "TEMPERATURAS":           ("dbo",     "Temperaturas por ciudad", "Horaria"),
    "trades":                 ("trading", "Registro de operaciones de trading ejecutadas", "Por operacion"),
}

for i, (tbl, (schema, desc, freq)) in enumerate(TABLE_META.items(), start=5):
    key = f"{schema}.{tbl}"
    cnt = row_counts.get(key, "N/A")
    dr  = date_ranges.get(tbl, ("-", "-"))
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    for col, val in enumerate([schema, tbl, desc, cnt, dr[0], dr[1], freq], start=1):
        c = ws1.cell(row=i, column=col, value=val)
        c.fill = fill; c.font = NORM_FONT; c.border = BORDER
        c.alignment = CENTER if col in (1, 4, 5, 6, 7) else LEFT

ws1.column_dimensions["A"].width = 10
ws1.column_dimensions["B"].width = 24
ws1.column_dimensions["C"].width = 48
ws1.column_dimensions["D"].width = 12
ws1.column_dimensions["E"].width = 14
ws1.column_dimensions["F"].width = 14
ws1.column_dimensions["G"].width = 14

# ── Hoja 2: Detalle de Columnas ───────────────────────────────────────────────
ws2 = wb.create_sheet("Columnas por Tabla")

ws2.merge_cells("A1:G1")
ws2["A1"] = "XTS Platform — Detalle de Columnas por Tabla"
ws2["A1"].font = TITLE_FONT
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30

for col, h in enumerate(
    ["Schema", "Tabla", "Columna", "Tipo", "Longitud", "Nullable", "PK"],
    start=1,
):
    c = ws2.cell(row=3, column=col, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
ws2.row_dimensions[3].height = 22

prev_table = None
row_num = 4
for _, r in df_schema.iterrows():
    tbl_key = f"{r.TABLE_SCHEMA}.{r.TABLE_NAME}"
    if prev_table != tbl_key:
        fill = GREEN_FILL
    else:
        fill = ALT_FILL if row_num % 2 == 0 else WHITE_FILL
    vals = [
        r.TABLE_SCHEMA, r.TABLE_NAME, r.COLUMN_NAME,
        r.DATA_TYPE.upper(),
        str(r.CHARACTER_MAXIMUM_LENGTH) if pd.notna(r.CHARACTER_MAXIMUM_LENGTH) else "-",
        r.IS_NULLABLE,
        r.IS_PK,
    ]
    for col, val in enumerate(vals, start=1):
        c = ws2.cell(row=row_num, column=col, value=val)
        c.fill = fill
        c.font = BOLD_FONT if r.IS_PK == "SI" else NORM_FONT
        c.border = BORDER
        c.alignment = CENTER if col in (1, 4, 5, 6, 7) else LEFT
    prev_table = tbl_key
    row_num += 1

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 24
ws2.column_dimensions["C"].width = 24
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 10
ws2.column_dimensions["F"].width = 10
ws2.column_dimensions["G"].width = 6

# ── Hoja 3: Glosario de Campos ────────────────────────────────────────────────
ws3 = wb.create_sheet("Glosario de Campos")

ws3.merge_cells("A1:D1")
ws3["A1"] = "Glosario de Campos Clave"
ws3["A1"].font = TITLE_FONT
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 30

for col, h in enumerate(["Campo", "Tabla", "Descripcion", "Unidad"], start=1):
    c = ws3.cell(row=3, column=col, value=h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER

GLOSSARY = [
    ("DA_DCL",    "DATOS_ERCOT", "Precio Day-Ahead nodo DC_L (Laredo, frontera Mexico-Texas)", "USD/MWh"),
    ("DA_DCR",    "DATOS_ERCOT", "Precio Day-Ahead nodo DC_R (Roma, frontera Mexico-Texas)",   "USD/MWh"),
    ("RT_DCL",    "DATOS_ERCOT", "Precio Real-Time nodo DC_L (Laredo)",                        "USD/MWh"),
    ("RT_DCR",    "DATOS_ERCOT", "Precio Real-Time nodo DC_R (Roma)",                          "USD/MWh"),
    ("SOLAR_ERCOT","DATOS_ERCOT","Generacion solar total del sistema ERCOT (system-wide)",     "GWh"),
    ("PML_IVY",   "DATOS_CAISO", "Precio Marginal Local nodo Ivyglen (CAISO, California)",    "USD/MWh"),
    ("PML_OMS",   "DATOS_CAISO", "Precio Marginal Local nodo Omslna (CAISO, California)",     "USD/MWh"),
    ("PPOE",      "GTM",         "Precio Programado del Organismo Ejecutor — forecast del despacho diario (AMM Guatemala)", "USD/MWh"),
    ("POE",       "GTM",         "Precio del Organismo Ejecutor — precio real/spot del post-despacho (AMM Guatemala)",     "USD/MWh"),
    ("LBR",       "GTM",         "PML nodo 09LBR-230 CENACE SIN — nodo de interconexion Mexico-Guatemala", "USD/MWh"),
    ("FIX",       "TIPO_CAMBIO", "Tipo de cambio FIX publicado por Banxico",                  "MXN/USD"),
    ("Archivo",   "ESTADOS_DE_CUENTA_XML", "Nombre del archivo XML del estado de cuenta",     "-"),
    ("FechaEDC",  "ESTADOS_DE_CUENTA_XML", "Fecha de liquidacion del estado de cuenta",       "DATE"),
    ("CO",        "ESTADOS_DE_CUENTA_XML", "Codigo de operacion / subcuenta CENACE",          "-"),
    ("Monto",     "ESTADOS_DE_CUENTA_XML", "Monto de la transaccion",                         "MXN"),
    ("TotalNeto", "ESTADOS_DE_CUENTA_XML", "Total neto del estado de cuenta",                 "MXN"),
    ("mw_max",    "CAPACIDADES",           "Capacidad maxima de transmision entre nodos",      "MW"),
    ("mw_excedente", "DEF_EXC",            "Energia excedente en el periodo",                 "MW"),
    ("mw_deficiente","DEF_EXC",            "Energia deficiente en el periodo",                "MW"),
    ("precio_da", "trading.trades",        "Precio Day-Ahead al que se ejecuto la operacion", "USD/MWh"),
    ("precio_rt", "trading.trades",        "Precio Real-Time de liquidacion de la operacion", "USD/MWh"),
]

for i, (campo, tabla, desc, unidad) in enumerate(GLOSSARY, start=4):
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    for col, val in enumerate([campo, tabla, desc, unidad], start=1):
        c = ws3.cell(row=i, column=col, value=val)
        c.fill = fill; c.font = NORM_FONT; c.border = BORDER
        c.alignment = LEFT if col == 3 else CENTER

ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 24
ws3.column_dimensions["C"].width = 62
ws3.column_dimensions["D"].width = 12

# ── Guardar ───────────────────────────────────────────────────────────────────
output = r"C:\Users\xiixt\Downloads\XTS_Estructura_BaseDatos.xlsx"
wb.save(output)
print(f"Excel guardado: {output}")
